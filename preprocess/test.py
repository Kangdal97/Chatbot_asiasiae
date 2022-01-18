import openpyxl   ### 엑셀 다루기 위한 패키지
import random
from openpyxl import Workbook, load_workbook
# from kobert_transformers import get_tokenizer
from kogpt2_transformers import get_kogpt2_tokenizer


########################################### step 1.  ###########################################
# root_path = "../data"
# tweet_file = root_path + "/tweeter_dialog_dataset.xlsx"
# tweet_file_output = root_path + "/tweeter_dialog_data.txt"
#
# f = open(tweet_file_output, 'w', encoding='utf-8')
#
# wb = load_workbook(filename=tweet_file)
#
# ws = wb[wb.sheetnames[0]]
# # print(sheet)
# for row in ws.iter_rows():
#   for cell in row:
#     if cell.value == None:
#       break
#     print(cell.value)
#     f.write(cell.value + "\n")
#   print("\n\n\n")
#   f.write("\n\n\n")
#
# f.close()

########################################### step 2.  ###########################################
# # def tweet_data_for_autoregressive():
# root_path = "../data"
#
# # wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
# # wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
#
# file_path = root_path + "/tweeter_dialog_data.txt"
# tweeter_autoregressive_file = root_path + "/tweeter_dialog_for_autoregressive.txt"
#
# data_file = open(file_path, 'r', encoding='utf-8')
# tweet_file = open(tweeter_autoregressive_file, 'w', encoding='utf-8')
#
# data_file_lines = data_file.readlines()
# dialog = ''
# for line_num, line_data in enumerate(data_file_lines):
#   if line_data == "\n" and dialog != '':
#     dialog += "\n"
#     tweet_file.write(dialog)
#     print(dialog)
#     dialog = ''
#   elif line_data != "\n":
#     dialog += "<s>" + line_data[:-1] + "</s>"
# data_file.close()
# tweet_file.close()


########################################### step 1-1.  wellness_dialog_dataset.xlsx  ###########################################
# def wellness_question_data():
#
# root_path = "../data"
# wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
# wellness_q_output = root_path + "/wellness_dialog_question.txt"
#
# f = open(wellness_q_output, 'w', encoding='utf-8')
#
# wb = load_workbook(filename=wellness_file)
#
# ws = wb[wb.sheetnames[0]]
# # print(sheet)
# for row in ws.iter_rows():
#     f.write(row[0].value + "    " + row[1].value + "\n")
#
# f.close()

########################################### step 1-2.  wellness_dialog_dataset.xlsx  ###########################################
# def wellness_answer_data():
# root_path = "../data"
# wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
# wellness_a_output = root_path + "/wellness_dialog_answer.txt"
#
# f = open(wellness_a_output, 'w', encoding='utf-8')
# wb = load_workbook(filename=wellness_file)
# ws = wb[wb.sheetnames[0]]
#
# for row in ws.iter_rows():
#     if row[2].value == None:
#       continue
#     else:
#       f.write(row[0].value + "    " + row[2].value + "\n")
# f.close()

########################################### step 1-3.  wellness_dialog_dataset.xlsx  ###########################################
# def category_data():
# root_path = "../data"
# data_path = root_path + "/wellness_dialog_dataset.txt"
# c_output = root_path + "/wellness_dialog_category.txt"
#
# i_f = open(data_path, 'r', encoding='utf-8')
# o_f = open(c_output, 'w', encoding='utf-8')
#
# category_count = 0
# flag = True
#
# cate_dict = []
# i_lines = i_f.readlines()
# for i, data in enumerate(i_lines):
#     tmp = data.split(',')
#     # a = tmp[1]
#     q = tmp[0]
#     if q not in cate_dict and q != '구분':
#         cate_dict.append(q)
#         o_f.write(q.strip() + "    " + str(category_count) + "\n")
#         category_count += 1
# o_f.close()
# i_f.close()


########################################### step 1-4.  wellness_dialog_for_text_classification.txt  ###########################################
# def wellness_text_classification_data():
# root_path = "../data"
# wellness_category_file = root_path + "/wellness_dialog_category.txt"
# wellness_question_file = root_path + "/wellness_dialog_question.txt"
# wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
#
# cate_file = open(wellness_category_file, 'r', encoding='utf-8')
# ques_file = open(wellness_question_file, 'r', encoding='utf-8')
# text_classfi_file = open(wellness_text_classification_file, 'w', encoding='utf-8')
#
# category_lines = cate_file.readlines()
# cate_dict = {}
# for line_num, line_data in enumerate(category_lines):
#     data = line_data.split('    ')
#     cate_dict[data[0]] = data[1][:-1]
# print(cate_dict)
#
# ques_lines = ques_file.readlines()
# ques_dict = {}
# for line_num, line_data in enumerate(ques_lines):
#     data = line_data.split('    ')
#     # print(data[1]+ "    " + cate_dict[data[0]])
#     text_classfi_file.write(data[1][:-1] + "    " + cate_dict[data[0]] + "\n")
#
# cate_file.close()
# ques_file.close()
# text_classfi_file.close()

########################################### step 1-5.  wellness_dialog_for_text_classification.txt  ###########################################
# def wellness_dialog_for_autoregressive():
# root_path = "../data"
# wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
# wellness_answer_file = root_path + "/wellness_dialog_answer.txt"
# wellness_question_file = root_path + "/wellness_dialog_question.txt"
# wellness_autoregressive_file = root_path + "/wellness_dialog_for_autoregressive_2.txt"
#
#
# answ_file = open(wellness_answer_file, 'r', encoding='utf-8')
# ques_file = open(wellness_question_file, 'r', encoding='utf-8')
# autoregressive_file = open(wellness_autoregressive_file, 'w', encoding='utf-8')
#
# answ_lines = answ_file.readlines()
# ques_lines = ques_file.readlines()
# ques_dict = {}
# for line_num, line_data in enumerate(ques_lines):
#     ques_data = line_data.split('    ')
#     for ans_line_num, ans_line_data in enumerate(answ_lines):
#         ans_data = ans_line_data.split('    ')
#         if ques_data[0] == ans_data[0]:
#             autoregressive_file.write(ques_data[1][:-1] + "    " + ans_data[1])
#         else:
#             continue
#
# answ_file.close()
# ques_file.close()
# autoregressive_file.close()

########################################### step 1-6.  tweeter_dialog_for_autoregressive.txt  ###########################################
# def tweet_data_for_autoregressive():
# root_path = "../data"
#
# # wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
# # wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
# file_path = root_path + "/tweeter_dialog_data.txt"
# tweeter_autoregressive_file = root_path + "/tweeter_dialog_for_autoregressive.txt"
#
# data_file = open(file_path, 'r', encoding='utf-8')
# tweet_file = open(tweeter_autoregressive_file, 'w', encoding='utf-8')
#
# data_file_lines = data_file.readlines()
# dialog = ''
# for line_num, line_data in enumerate(data_file_lines):
#   if line_data == "\n" and dialog != '':
#       dialog += "\n"
#       tweet_file.write(dialog)
#       print(dialog)
#       dialog = ''
#   elif line_data != "\n":
#       dialog += "<s>" + line_data[:-1] + "</s>"
# data_file.close()
# tweet_file.close()


########################################### step 1-7.  tweeter_dialog_for_autoregressive.txt  ###########################################
# def seperate_wellness_data():
# root_path = "../data"
# # wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
# # wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
# file_path = root_path + "/wellness_dialog_for_autoregressive_2.txt"
# train_file_path = root_path + "/wellness_dialog_for_autoregressive_train.txt"
# test_file_path = root_path + "/wellness_dialog_for_autoregressive_test.txt"
#
# sperated_file = open(file_path, 'r', encoding='utf-8')
# train_file = open(train_file_path, 'w', encoding='utf-8')
# test_file = open(test_file_path, 'w', encoding='utf-8')
#
# sperated_file_lines = sperated_file.readlines()
# ques_dict = {}
# for line_num, line_data in enumerate(sperated_file_lines):
#     rand_num = random.randint(0, 10)
#     if rand_num < 10:
#         train_file.write(line_data)
#     else:
#         test_file.write(line_data)
#
# sperated_file.close()
# train_file.close()
# test_file.close()

########################################### step 1-8.  tweeter_dialog_for_autoregressive_token.txt  ###########################################
# # def tweeter_autoregressive_data():
# root_path = "../data"
# tokenizer =get_kogpt2_tokenizer()
# # wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
# # wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
# file_path = root_path + "/tweeter_dialog_for_autoregressive.txt"
# tweeter_autoregressive_file = root_path + "/tweeter_dialog_for_autoregressive_with_token.txt"
#
# data_file = open(file_path, 'r', encoding='utf-8')
# tweet_file = open(tweeter_autoregressive_file, 'w', encoding='utf-8')
#
# data_file_lines = data_file.readlines()
# dialog = ''
# max_len=0
# for line_num, line_data in enumerate(data_file_lines):
#     if line_data == "\n" and dialog != '':
#         dialog += "\n"
#         tweet_file.write(dialog)
#         print(dialog)
#         dialog = ''
#     elif line_data != "\n":
#         tmp_data = dialog + "<s>" + line_data[:-1] + "</s>"
#         if len(tokenizer.encode(tmp_data))>= 1024:
#             continue
#         else:
#             max_len= max(len(tokenizer.encode(tmp_data)),max_len)
#             dialog = tmp_data
# print('max_token_length: ', max_len)
# data_file.close()
# tweet_file.close()



########################################### step 1-10.  wellness_dialog_for_autoregressive_with_token.txt  ###########################################
# def tweeter_autoregressive_data_with_token():

root_path = "../data"
wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
wellness_answer_file = root_path + "/wellness_dialog_answer.txt"
wellness_question_file = root_path + "/wellness_dialog_question.txt"
wellness_autoregressive_file = root_path + "/wellness_dialog_for_autoregressive_with_token.txt"

answ_file = open(wellness_answer_file, 'r', encoding='utf-8')
ques_file = open(wellness_question_file, 'r', encoding='utf-8')
autoregressive_file = open(wellness_autoregressive_file, 'w', encoding='utf-8')

answ_lines = answ_file.readlines()
ques_lines = ques_file.readlines()
ques_dict = {}
for line_num, line_data in enumerate(ques_lines):
    ques_data = line_data.split('    ')
    for ans_line_num, ans_line_data in enumerate(answ_lines):
        ans_data = ans_line_data.split('    ')
        if ques_data[0] == ans_data[0]:
            autoregressive_file.write("<s>" + ques_data[1][:-1] + "</s><s>" + ans_data[1][:-1] + "</s>\n")
        else:
            continue

answ_file.close()
ques_file.close()
autoregressive_file.close()